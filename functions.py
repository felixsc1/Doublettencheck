import pandas as pd
import os
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_reference_ids(folder_path):
    """
    Extract ReferenceID from all sheets of all Excel files in the specified folder and its subfolders.
    """
    reference_ids = set()

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                try:
                    # Open the Excel file
                    with pd.ExcelFile(file_path, engine='openpyxl') as xls:
                        # Iterate over all sheets
                        for sheet_name in xls.sheet_names:
                            df = pd.read_excel(xls, sheet_name=sheet_name)
                            if 'ReferenceID' in df.columns:
                                reference_ids.update(df['ReferenceID'].dropna().unique())
                except Exception as e:
                    print(f"Error reading {file_path}: {e}")

    return reference_ids


def highlight_matches_in_excel(file_path, reference_ids, typ):
    # Read the Excel file
    df = pd.read_excel(file_path, engine='openpyxl')

    # Filter the DataFrame by including rows where the "Typ" column matches any of the entries in the list
    df = df[df['Typ'].isin(typ)]

    # Counters
    match_count = 0
    total_count = len(df)

    # Check if the ReferenceID column exists
    if 'ReferenceID' not in df.columns:
        print("No 'ReferenceID' column found in the file.")
        return

    # Split the file_path to add the suffix
    dir_name, file_name = os.path.split(file_path)
    file_base, file_extension = os.path.splitext(file_name)
    output_file_name = f"{file_base}_doublettencheck{file_extension}"
    output_path = os.path.join(dir_name, output_file_name)

    # Create an Excel writer with openpyxl engine
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Create a green fill
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

        # Iterate through the DataFrame rows
        for row_index, row in df.iterrows():
            if row['ReferenceID'] in reference_ids:
                match_count += 1
                for col_index in range(len(row) + 1):
                    worksheet.cell(row=row_index + 2, column=col_index + 1).fill = green_fill

    # Save the workbook
    workbook.save(output_path)

    # Print the number of matches and total length
    print(f"Total rows: {total_count}, Matching rows: {match_count}")
    print(f"{match_count / total_count * 100:.2f}% of Doubletten found.")



def main():
    folder_path = '~/GitRepos/BAKOM/output/organisationsdoubletten_auswertung'
    folder_path = os.path.expanduser(folder_path)
    ids = extract_reference_ids(folder_path)

    total_doubletten_file = '~/GitRepos/BAKOM/data/doublettencheck/012 - Liste Doubletten - BAKOM.xlsx'
    total_doubletten_file = os.path.expanduser(total_doubletten_file)
    highlight_matches_in_excel(total_doubletten_file, ids)

if __name__ == "__main__":
    main()
